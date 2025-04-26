from openpyxl import Workbook, load_workbook 

def _load_database() -> Workbook:
    return load_workbook("./back/data/companies.xlsx")

def save_database(wb: Workbook):
    wb.save("./back/data/companies.xlsx")

def get_companies() -> list[str]:
    wb = _load_database()
    ws = wb.active
    companies = []
    for col in ws.iter_cols(min_col=0, max_col=ws.max_column):
        if col[0].value.lower() in ["domain", "domains"]:
            for row in col[1:]:
                companies.append(row.value)
    
    return companies

def update_info(domain: str, data: dict):
    try:
        domain_idx = None
        
        wb = _load_database()
        ws = wb.active

        print("Loaded ws, iterating through cols")
        # Iterate through the first column to find the domain row
        for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
            if col[0].value and col[0].value.lower() in ["domain", "domains"]:
                print(f"Found domain column, looking for domain {domain}")
                # Search for the domain in this column
                for row_idx, row in enumerate(col[1:], start=1):  # Skip the header
                    if row.value == domain:
                        domain_idx = row_idx
                        print(f"Domain found at row {domain_idx}")
                        break

        # If the domain index was found, proceed with updating values
        if domain_idx is not None:
            # Iterate over the columns to find "id" and "name" columns
            for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
                header_value = col[0].value
                if header_value and header_value.lower() in ["id", "ids"]:
                    print(f"Found id column, updating value for domain {domain}")
                    try:
                        col[domain_idx].value = data["organization"]["id"]
                    except Exception as e:
                        print(f"Failed to set id: {e}")
                
                if header_value and header_value.lower() in ["name", "names", "denomination", "denominations"]:
                    print(f"Found name column, updating value for domain {domain}")
                    try:
                        col[domain_idx].value = data["organization"]["name"]
                    except Exception as e:
                        print(f"Failed to set name: {e}")
            
            # Save the workbook after making changes
            save_database(wb)
            print(f"Updated data for {domain} successfully!")

        else:
            print(f"Domain {domain} not found!")
            raise Exception("Domain not found!")

    except Exception as e:
        print(f"[ERROR @ excel_manager/update_info()] {e}")